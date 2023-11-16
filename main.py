import openpyxl
from scheldue import Scheldue
wb = openpyxl.load_workbook('16.11.2023.xlsx')

ws = wb.active
skip_arr = ["1 пара", "2 пара", "3 пара", "4 пара", "5 пара"]
print(f'Total number of rows: {ws.max_row}. And total number of columb: {ws.max_column}')

data_cabinet = []
data_teacher = []
data_lesson = []
for i in range(1, ws.max_row - 1):
    for j in range(1, ws.max_column - 1):
        val = ws.cell(row=i, column=j).value
        if val is None or str(val).strip() in skip_arr or str(val).strip() == "":
            continue
        else:
            if "/" in str(val):
                data_cell = str(val).split("/")
                data_lesson.append(data_cell[0])
                data_teacher.append(data_cell[1])
            if len(str(val)) == 3 or len(str(val)) == 4 or val == "Академия КП":
                data_cabinet.append(val)

data_classes = []
for all_i in range(len(data_cabinet)):
    obj = Scheldue(data_lesson[all_i], data_teacher[all_i], data_cabinet[all_i])
    data_classes.append(obj)
    
cab = input("Кабинет: ")
for user in data_classes:
    if cab in str(user.cabinet):
        print(f"{user.teacher} in {user.cabinet}")

